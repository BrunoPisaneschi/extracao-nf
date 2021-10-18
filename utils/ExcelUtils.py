from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from os import path, makedirs
from datetime import datetime



class excel():
    def __init__(self, file_path=".\\relatorio\\",
                 file_name=datetime.today().strftime("%m-%Y") + " - ExtracaoNFs", file_extension=".xlsx"):
        '''When you create Instance of excel, you can change file_path, file_name and file_extension'''
        self._wb = None
        self._ws = None
        self._file_path = file_path
        self._file_name = file_name
        self._file_extension = file_extension
        self._dict_datas = {}

    def _configure_header(self):
        font = Font(color="FFFFFF", bold=True, size=12)
        fill = PatternFill("solid", fgColor="357ae8")
        alignment = Alignment(horizontal="center", vertical="center")
        self._ws.cell(row=1, column=1).font = font
        self._ws.cell(row=1, column=2).font = font
        self._ws.cell(row=1, column=3).font = font
        self._ws.cell(row=1, column=4).font = font
        self._ws.cell(row=1, column=1).fill = fill
        self._ws.cell(row=1, column=2).fill = fill
        self._ws.cell(row=1, column=3).fill = fill
        self._ws.cell(row=1, column=4).fill = fill
        self._ws.cell(row=1, column=1).alignment = alignment
        self._ws.cell(row=1, column=2).alignment = alignment
        self._ws.cell(row=1, column=3).alignment = alignment
        self._ws.cell(row=1, column=4).alignment = alignment

    def write_excel(self, dict_datas):
        '''Write Excel receive a dictionary'''
        self._dict_datas = dict_datas
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.cell(row=1, column=1).value = 'Nome'
        self._ws.cell(row=1, column=2).value = 'Mês vigente'
        self._ws.cell(row=1, column=3).value = 'Valor'
        self._ws.cell(row=1, column=4).value = 'link pdf'
        self._configure_header()

        for _index, _data in self._dict_datas.items():
            self._ws.cell(row=int(_index+1), column=1).value = _data['nome']
            self._ws.cell(row=int(_index+1), column=2).value = _data['mes_vigente']
            self._ws.cell(row=int(_index+1), column=3).value = _data['valor']
            self._ws.cell(row=int(_index+1), column=4).value = _data['link_pdf']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._ws.title = 'NFs'
        self._wb.save(self._file_path + self._file_name + self._file_extension)
        return self._file_path + self._file_name + self._file_extension

    def update_excel(self, dict_datas):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _max_row = self._ws.max_row

        self._ws.title = 'NFs'
        for _index, _data in self._dict_datas.items():
            self._ws.cell(row=int(_index+_max_row), column=1).value = _data['nome']
            self._ws.cell(row=int(_index+_max_row), column=2).value = _data['mes_vigente']
            self._ws.cell(row=int(_index+_max_row), column=3).value = _data['valor']
            self._ws.cell(row=int(_index+_max_row), column=4).value = _data['link_pdf']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._wb.save(self._file_path + self._file_name + self._file_extension)
        self._wb.close()
        return self._file_path + self._file_name + self._file_extension
